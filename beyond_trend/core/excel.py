from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def resolve_attr(obj, path):
    """Walk a dotted/double-underscore attribute path, calling no-arg callables."""
    for part in path.split("__"):
        if obj is None:
            return ""
        obj = getattr(obj, part, "")
        if callable(obj):
            obj = obj()
    return obj


def _to_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value).replace(tzinfo=None)
        return value
    if isinstance(value, (date, int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def build_workbook(sheets):
    """Build an openpyxl Workbook from a list of sheet dicts.

    Each sheet dict: {"name": str, "headers": list[str], "rows": list[list]}.
    """
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"][:31])
        headers = sheet["headers"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for row in sheet["rows"]:
            ws.append([_to_cell(v) for v in row])
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20
    return wb


def workbook_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ExcelExportMixin:
    """Admin mixin that adds an `export_to_excel` action.

    Subclasses should set ``excel_export_fields`` as a list of
    ``(header_label, attribute_path)`` tuples. Override
    ``get_excel_sheets`` to emit additional sheets (e.g. line items).
    """

    excel_export_fields: list = []
    excel_sheet_name: str | None = None
    excel_filename_prefix: str | None = None

    def get_excel_sheets(self, request, queryset):
        headers = [label for label, _ in self.excel_export_fields]
        rows = [
            [resolve_attr(obj, path) for _, path in self.excel_export_fields]
            for obj in queryset
        ]
        name = self.excel_sheet_name or self.model._meta.verbose_name_plural.title()
        return [{"name": name, "headers": headers, "rows": rows}]

    def export_to_excel(self, request, queryset):
        sheets = self.get_excel_sheets(request, queryset)
        wb = build_workbook(sheets)
        prefix = self.excel_filename_prefix or self.model._meta.model_name
        filename = f"{prefix}_{timezone.now():%Y%m%d_%H%M%S}.xlsx"
        return workbook_response(wb, filename)

    export_to_excel.short_description = "Export selected to Excel"
